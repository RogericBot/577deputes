"""Population (INSEE) + inscrits/votants (Min. Intérieur) per circonscription.

Two sources merged on (dept_code, circo_num) :

  * INSEE — Populations légales 2021 par circonscription, XLSX 32 KB.
    Mise en ligne 25/06/2024, mise à jour annuelle.
    https://www.insee.fr/fr/statistiques/fichier/2508230/
        population-circonscriptions-legislatives-2021.xlsx

  * Ministère de l'Intérieur — Législatives 2024 T1 (résultats définitifs),
    CSV 330 KB. Stable jusqu'à la prochaine élection législative
    (sauf élections partielles).
    https://static.data.gouv.fr/resources/elections-legislatives-des-30-juin-…
        /resultats-definitifs-par-circonscriptions-legislatives.csv

Les 11 circos des Français de l'étranger n'ont pas de population
résidente : `population` reste NULL pour ces lignes.
"""
from __future__ import annotations

import csv
import sqlite3
import time
from pathlib import Path

import httpx

from ..config import settings
from ..logging_setup import get_logger

log = get_logger(__name__)

INSEE_URL = (
    "https://www.insee.fr/fr/statistiques/fichier/2508230/"
    "population-circonscriptions-legislatives-2021.xlsx"
)
MININT_URL = (
    "https://static.data.gouv.fr/resources/"
    "elections-legislatives-des-30-juin-et-7-juillet-2024-resultats-definitifs-"
    "du-1er-tour/20240710-171413/"
    "resultats-definitifs-par-circonscriptions-legislatives.csv"
)


def _normalise_dept(value: str | None) -> str | None:
    """Return the INSEE dept code padded to 2 chars (`1` → `01`).
    Keeps `2A`/`2B`/`971`/etc. unchanged."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if s.isdigit():
        return s.zfill(2) if len(s) <= 2 else s
    return s


def _circo_num_from_code(code: str | None) -> int | None:
    """The Min. Intérieur encodes circo as `<dept_code><circo_num>` :
    `101` for dept 1 / circo 1, `7518` for Paris / circo 18.
    We extract the last 2 chars and parse to int."""
    if code is None:
        return None
    s = str(code).strip()
    if len(s) < 2:
        return None
    try:
        return int(s[-2:])
    except ValueError:
        return None


def _to_int(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip().replace(" ", "").replace(" ", "").replace("\xa0", "")
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return None


# ---------------------------------------------------------------------
# Downloads
# ---------------------------------------------------------------------
def download_files() -> tuple[Path, Path]:
    """Fetch both files into data/raw/, with simple ETag-aware caching."""
    insee_path = settings.raw_dir / "population_circo.xlsx"
    minint_path = settings.raw_dir / "legislatives_2024_t1.csv"
    started = time.time()
    with httpx.Client(timeout=120.0, follow_redirects=True,
                      headers={"User-Agent": settings.user_agent}) as client:
        for url, dst in ((INSEE_URL, insee_path), (MININT_URL, minint_path)):
            r = client.get(url)
            r.raise_for_status()
            dst.write_bytes(r.content)
            log.info("circo_stats_download", extra={"url": url, "bytes": len(r.content)})
    log.info("circo_stats_downloaded", extra={"elapsed_s": round(time.time() - started, 1)})
    return insee_path, minint_path


# ---------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------
def parse_insee_population(xlsx_path: Path) -> dict[tuple[str, int], int]:
    """Return {(dept_code, circo_num): population}."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Population par circonscription"]
    out: dict[tuple[str, int], int] = {}
    headers: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None for c in row):
            continue
        # Skip until we hit the header row that starts with "Code Département".
        if headers is None:
            if row and isinstance(row[0], str) and "Code" in row[0] and "épartement" in row[0]:
                headers = [str(c) if c else "" for c in row]
            continue
        # Data rows : col0 = dept, col1 = circo, col3 = population
        dept = _normalise_dept(row[0])
        circo = _circo_num_from_code(row[1])
        pop = _to_int(row[3])
        if dept is None or circo is None or pop is None:
            continue
        out[(dept, circo)] = pop
    return out


def parse_minint_inscrits(csv_path: Path) -> dict[tuple[str, int], dict[str, int]]:
    """Return {(dept_code, circo_num): {inscrits, votants, abstentions}}."""
    out: dict[tuple[str, int], dict[str, int]] = {}
    with csv_path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            dept = _normalise_dept(row.get("Code département"))
            circo = _circo_num_from_code(row.get("Code circonscription législative"))
            if dept is None or circo is None:
                continue
            out[(dept, circo)] = {
                "inscrits": _to_int(row.get("Inscrits")),
                "votants": _to_int(row.get("Votants")),
                "abstentions": _to_int(row.get("Abstentions")),
            }
    return out


# ---------------------------------------------------------------------
# Bulk loader
# ---------------------------------------------------------------------
def ingest_circo_stats(conn: sqlite3.Connection) -> dict[str, int]:
    """End-to-end : download (or use cached) + parse + upsert."""
    insee_path = settings.raw_dir / "population_circo.xlsx"
    minint_path = settings.raw_dir / "legislatives_2024_t1.csv"
    if not insee_path.exists() or not minint_path.exists():
        download_files()

    log.info("circo_stats_parse_start")
    pop = parse_insee_population(insee_path)
    inscr = parse_minint_inscrits(minint_path)

    # Union of keys.
    keys = set(pop) | set(inscr)
    rows: list[tuple] = []
    for (dept, circo) in keys:
        d_inscr = inscr.get((dept, circo)) or {}
        rows.append((
            dept, circo,
            pop.get((dept, circo)),
            d_inscr.get("inscrits"),
            d_inscr.get("votants"),
            d_inscr.get("abstentions"),
            INSEE_URL,
            MININT_URL,
        ))

    conn.execute("BEGIN")
    conn.execute("DELETE FROM circo_stats")
    conn.executemany(
        "INSERT INTO circo_stats(dept_code, circo_num, population, inscrits, "
        "votants, abstentions, source_pop, source_inscr) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    conn.execute("COMMIT")

    n = conn.execute("SELECT COUNT(*) AS c FROM circo_stats").fetchone()["c"]
    n_pop = conn.execute(
        "SELECT COUNT(*) AS c FROM circo_stats WHERE population IS NOT NULL"
    ).fetchone()["c"]
    n_inscr = conn.execute(
        "SELECT COUNT(*) AS c FROM circo_stats WHERE inscrits IS NOT NULL"
    ).fetchone()["c"]
    log.info(
        "circo_stats_done",
        extra={"rows": n, "with_population": n_pop, "with_inscrits": n_inscr},
    )
    return {"rows": n, "with_population": n_pop, "with_inscrits": n_inscr}
